import os
import pandas as pd
import numpy as np
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def create_room_schedule(room_data):
    """Creates a schedule DataFrame for a single room"""
    time_blocks = {
        1: "08:00 - 09:00",
        2: "09:15 - 10:15",
        3: "10:30 - 11:30",
        4: "11:45 - 12:45",
        5: "12:50 - 13:50",
        6: "13:55 - 14:55",
        7: "15:00 - 16:00",
        8: "16:15 - 17:15",
        9: "17:30 - 18:30",
    }
    days = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
    
    schedule_df = pd.DataFrame(
        index=time_blocks.values(),
        columns=days
    )
    
    # Fill schedule with room's subjects
    for subject in room_data['Asignaturas']:
        time_slot = time_blocks[subject['Bloque']]
        day = subject['Dia']
        content = (f"Asignatura: {subject['Nombre']}\n"
                  f"Satisfacción: {subject['Satisfaccion']}/10")
        schedule_df.at[time_slot, day] = content
    
    return schedule_df.fillna('')

def apply_excel_styling(worksheet):
    """Applies styling to the Excel worksheet"""
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFCCE5', end_color='FFCCE5', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in worksheet.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:  # Headers
                cell.fill = header_fill
                cell.font = Font(bold=True)

def save_room_schedules(data, filename='room_schedules.xlsx'):
    """Saves each room's schedule to a separate worksheet"""

    # Create scheduleRepresentation directory if it doesn't exist
    output_dir = 'scheduleRepresentation'
    os.makedirs(output_dir, exist_ok=True)
    
    # Construct full file path
    filepath = os.path.join(output_dir, filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for room_data in data:
            room_code = room_data['Codigo']
            capacity = room_data['Capacidad']
            schedule_df = create_room_schedule(room_data)
            
            # Write DataFrame to Excel
            sheet_name = f"Sala {room_code}"[:31]  # Excel sheet names limited to 31 chars
            schedule_df.to_excel(writer, sheet_name=sheet_name, index=True)
            
            # Apply styling
            worksheet = writer.sheets[sheet_name]
            apply_excel_styling(worksheet)
            
            # Add room capacity info
            capacity_cell = worksheet.cell(row=1, column=1)
            capacity_cell.value = f"Capacidad: {capacity} estudiantes"
            
            # Adjust column widths
            for idx, col in enumerate(schedule_df.columns):
                worksheet.column_dimensions[chr(66 + idx)].width = 30  # B, C, D, E, F
            worksheet.column_dimensions['A'].width = 15  # Time column
            
            # Set row heights
            for row in range(1, len(schedule_df) + 2):
                worksheet.row_dimensions[row].height = 60

def main():
    try:
        with open("agent_output/Horarios_salas.json", 'r', encoding="utf-8") as file:
            schedule_data = json.load(file)
            save_room_schedules(schedule_data)
            print(f"Schedules generated successfully for {len(schedule_data)} rooms")
            
            # Print summary
            for room in schedule_data:
                print(f"\nRoom: {room['Codigo']}")
                print(f"Capacity: {room['Capacidad']} students")
                print(f"Subjects assigned: {len(room['Asignaturas'])}")
    except Exception as e:
        print(f"Error processing schedules: {str(e)}")

if __name__ == '__main__':
    main()